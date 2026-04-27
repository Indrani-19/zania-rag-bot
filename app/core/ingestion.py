import json
import logging
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader

from app.config import settings


logger = logging.getLogger(__name__)


class IngestionError(Exception):
    pass


class ScannedPdfError(IngestionError):
    pass


class EmptyPdfError(IngestionError):
    pass


# Caps array expansion so a single 1000-element array can't blow up into 1000 chunks.
JSON_ARRAY_EXPAND_LIMIT = 50


def load_pdf(content: bytes) -> list[tuple[int, str]]:
    reader = PdfReader(BytesIO(content))
    if len(reader.pages) == 0:
        raise EmptyPdfError("PDF has no pages.")

    pages: list[tuple[int, str]] = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append((i, text))

    if not any(text.strip() for _, text in pages):
        raise ScannedPdfError(
            f"PDF has {len(pages)} pages but no extractable text. "
            "It may be scanned/image-based — OCR is not supported."
        )

    return pages


def flatten_json(data: Any, prefix: str = "") -> list[str]:
    lines: list[str] = []

    if isinstance(data, dict):
        for key, value in data.items():
            new_prefix = f"{prefix}.{key}" if prefix else str(key)
            lines.extend(flatten_json(value, new_prefix))
    elif isinstance(data, list):
        if len(data) > JSON_ARRAY_EXPAND_LIMIT:
            for i, item in enumerate(data[:JSON_ARRAY_EXPAND_LIMIT]):
                lines.extend(flatten_json(item, f"{prefix}[{i}]"))
            lines.append(
                f"{prefix}: [array truncated — {len(data)} elements total, "
                f"only first {JSON_ARRAY_EXPAND_LIMIT} indexed]"
            )
        else:
            for i, item in enumerate(data):
                lines.extend(flatten_json(item, f"{prefix}[{i}]"))
    else:
        rendered = data if isinstance(data, str) else json.dumps(data)
        lines.append(f"{prefix}: {rendered}" if prefix else str(rendered))

    return lines


def load_json(content: bytes) -> str:
    try:
        data = json.loads(content)
    except json.JSONDecodeError as e:
        raise IngestionError(f"Invalid JSON: {e}") from e
    return "\n".join(flatten_json(data))


def load_xlsx(content: bytes) -> list[tuple[int, str]]:
    """Load .xlsx into (sheet_index, text) blocks. One sheet per "page".

    Each non-empty data row is rendered as a labeled key/value block using the
    first row as headers, e.g.:

        [Sheet1, row 2]
        question: Where are your data centres located?
        answer: Our data centers are located in the US Central region...

    Keeping rows together as units (rather than column-flattening) preserves
    Q&A semantics for retrieval — a question and its answer stay in the same chunk.
    """
    try:
        wb = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except (InvalidFileException, BadZipFile, KeyError, ValueError) as e:
        raise IngestionError(f"Invalid xlsx: {e}") from e

    if not wb.sheetnames:
        raise IngestionError("xlsx has no sheets")

    sheets: list[tuple[int, str]] = []
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        headers: list[str] = []
        for i, cell in enumerate(header_row, start=1):
            label = "" if cell is None else str(cell).strip()
            headers.append(label or f"col{i}")

        blocks: list[str] = []
        for row_idx, row in enumerate(rows_iter, start=2):
            kv_lines: list[str] = []
            for header, value in zip(headers, row):
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                kv_lines.append(f"{header}: {value}")
            if kv_lines:
                blocks.append(f"[{sheet_name}, row {row_idx}]\n" + "\n".join(kv_lines))

        if blocks:
            sheets.append((sheet_idx, "\n\n".join(blocks)))

    if not sheets:
        raise IngestionError("xlsx has no data rows")

    return sheets


def _splitter() -> RecursiveCharacterTextSplitter:
    return RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        length_function=len,
    )


def chunk_pdf_pages(pages: list[tuple[int, str]], source: str) -> list[Document]:
    # Chunk per-page so page numbers stay attached to each chunk's metadata.
    splitter = _splitter()
    documents: list[Document] = []
    for page_num, text in pages:
        if not text.strip():
            continue
        for chunk in splitter.split_text(text):
            documents.append(
                Document(
                    page_content=chunk,
                    metadata={"source": source, "page": page_num},
                )
            )
    return documents


def chunk_json_text(flattened: str, source: str) -> list[Document]:
    splitter = _splitter()
    return [
        Document(page_content=chunk, metadata={"source": source})
        for chunk in splitter.split_text(flattened)
        if chunk.strip()
    ]


def ingest(content: bytes, filename: str) -> list[Document]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        pages = load_pdf(content)
        documents = chunk_pdf_pages(pages, source=filename)
    elif suffix == ".json":
        flattened = load_json(content)
        documents = chunk_json_text(flattened, source=filename)
    elif suffix == ".xlsx":
        # Sheet index is reused as the "page" metadata so verbose responses
        # cite "Page 1" = "Sheet 1" for xlsx. chunk_pdf_pages is generic over
        # (int, text) pairs — the name is just historical.
        sheets = load_xlsx(content)
        documents = chunk_pdf_pages(sheets, source=filename)
    else:
        raise IngestionError(
            f"Unsupported file type: {suffix}. Supported: .pdf, .json, .xlsx"
        )

    logger.info(
        "Ingested %s: %d chunks (suffix=%s, bytes=%d)",
        filename, len(documents), suffix, len(content),
    )
    return documents
